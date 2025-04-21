from openpyxl import load_workbook
import datetime

def load_excel():
    wb = load_workbook('template.xlsx')
    sheet = wb.active
    lis = [[cell.value for cell in row] for row in sheet.iter_rows(min_row=2)] # создаем список начиная со 2-й ячейки excel
    if lis[-1][0] is None:  # убираем пустые ячейки в конце (хвост)
        del lis[-1]
    lis = list(map(lambda x: [*x[0].split(), x[1], x[2], x[3], x[4], x[5], x[6], x[7], x[8], x[9].strftime("%Y-%m-%d"), x[10]], lis)) # фио
    return lis

def xml_shablon(data):
    list_xml = []
    list_xml.append(f'<?xml version = "1.0" encoding = "utf-8"?>\n')
    list_xml.append(f'<RegistrySet>\n')
    for i in range(len(list_data)):
        list_xml.append(f' <RegistryRecord>\n')
        list_xml.append(f'  <Worker>\n')
        list_xml.append(f'   <LastName>{list_data[i][0]}</LastName>\n')
        list_xml.append(f'   <FirstName>{list_data[i][1]}</FirstName>\n')
        list_xml.append(f'   <MiddleName>{list_data[i][2]}</MiddleName>\n')
        list_xml.append(f'   <Snils>{list_data[i][3]}</Snils>\n')
        list_xml.append(f'   <Position>{list_data[i][4]}</Position>\n')
        list_xml.append(f'   <EmployerInn>{list_data[i][5]}</EmployerInn>\n')
        list_xml.append(f'   <EmployerTitle>{list_data[i][6]}</EmployerTitle>\n')
        list_xml.append(f'  </Worker>\n')
        list_xml.append(f'  <Organization>\n')
        list_xml.append(f'   <Inn>{list_data[i][7]}</Inn>\n')
        list_xml.append(f'   <Title>{list_data[i][8]}</Title>\n')
        list_xml.append(f'  </Organization>\n')
        list_xml.append(f'  <Test isPassed="true" learnProgramId="{list_data[i][-4][0]}">\n')
        list_xml.append(f'   <Date>{list_data[i][-2]}</Date>\n')
        list_xml.append(f'   <ProtocolNumber>{list_data[i][-1]}</ProtocolNumber>\n')
        list_xml.append(f'   <LearnProgramTitle>{list_data[i][9][2:]}</LearnProgramTitle>\n')
        list_xml.append(f'  </Test>\n')
        list_xml.append(f' </RegistryRecord>\n')
    list_xml.append(f'</RegistrySet>')
    #del list_xml[-1]
    #list_xml.append(f'      </RegistryRecord></RegistrySet>')
    return list_xml

list_data = load_excel() # данные из Excel - файла

print(list_data)

with open('filexml.xml', 'w', encoding='utf-8') as file:
    file.writelines(xml_shablon(list_data))

print(xml_shablon(list_data))